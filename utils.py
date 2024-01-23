from openpyxl import load_workbook
import numpy as np
import pprint

def get_columns(srcfile: str) -> (np.ndarray, dict):
    wb = load_workbook(srcfile)
    ws = wb['columns']
    columns = []
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        col = []
        columns.append(col)
        for cell in row:
            col.append(cell.value)
    columns = np.array(columns)
    print("Columns:")
    pprint.pprint(columns)
    students_groups = {}
    for ws in wb.worksheets:
        if ws.title != "columns":
            students_in_group = []
            students_groups[ws.title] = students_in_group
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
                student = []
                students_in_group.append(student)
                for cell in row:
                    student.append(cell.value)
            # print(f"Students in {ws.title}:")
            # pprint.pprint(students_in_group)
    return columns, students_groups

if __name__ == '__main__':
    pass