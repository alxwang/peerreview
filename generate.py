from argparse import ArgumentParser
from argparse import RawTextHelpFormatter
import numpy as np
from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from utils import *

def parse_arg():
    parser = ArgumentParser(description="Generate peer review files.", formatter_class=RawTextHelpFormatter,
                            epilog="Example：\n"
                                   "Excel：generate.py src.xlsx\n"
                            )
    parser.add_argument("src",
                        help="Read information from this excel file", metavar="Input-File", type=str)
    args = parser.parse_args()
    return vars(args)




if __name__ == '__main__':
    args = parse_arg()
    src = args["src"]
    print(f"Start generating peer review files from {src}...")

    cols, groups = get_columns(src)
    for group_name, students in groups.items():
        newwb = Workbook()
        newws = newwb.active
        newws.title = f"Peer Review for Team {group_name}"
        row = list(map(lambda x: f"{x[0]}({x[1]} to {x[2]})", cols.tolist()))
        row.insert(0, "Student Name")
        newws.append(row)
        for student in students:
            row = []
            row.append(f"{student[0]}, {student[1]}")
            for i in range(0, len(cols)):
                row.append(0)
            newws.append(row)

        for row in newws.iter_rows():
            for cell in row:
                if cell.col_idx == 1 or cell.row == 1:
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color='00CC00', end_color='00CC00', fill_type="solid")
                    cell.number_format = '0'

        newws.column_dimensions['A'].width = 30
        for col in newws.iter_cols(min_col=2, max_col=newws.max_column):
            col[0].alignment = Alignment(horizontal='center', vertical='center')
            newws.column_dimensions[get_column_letter(col[0].column)].width = 20

        newwb.save(f"Peer Review for Team {group_name}.xlsx")
        newwb.close()

    print("Complete generating peer review files...")
