from argparse import ArgumentParser
from argparse import RawTextHelpFormatter
import os
import sys
import platform
from pathlib import Path
import pprint
import numpy as np
from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
from utils import *
import os
import os.path
import pandas as pd

def parse_arg():
    parser = ArgumentParser(description="Collect peer review files.",formatter_class=RawTextHelpFormatter,
                            epilog=     "Example：\n"
                                        "Excel：collect.py folder src.xlsx dest.xlsx\n"
                                        )
    parser.add_argument("folder",
                        help="Read peer review excel files from this folder", metavar="Input-File", type=str)
    parser.add_argument("src",
                        help="Read information from this excel file", metavar="Input-File", type=str)
    parser.add_argument("dst",
                        help="Write results to this excel file", metavar="Output-File", type=str)

    args = parser.parse_args()
    return vars(args)

if __name__ == '__main__':
    args = parse_arg()
    src = args["src"]
    dst = args["dst"]
    folder = args["folder"]
    print(f"Start collecting peer review files from {folder} based on {src}...")
    print(f"Write results to {dst}...")
    print("Collecting peer review files...")

    cols, groups = get_columns(src)

    xlsx_files = []
    for dir_path, dir_names, file_names in os.walk(folder):
        for file_name in [f for f in file_names if f.endswith(".xlsx")]:
            xlsx_files.append(os.path.join(dir_path, file_name))

    row = list(map(lambda x: f"{x[0]}({x[1]} to {x[2]})", cols.tolist()))
    row.insert(0, "Student Name")

    #reviews[group_name] is a team's reviews
    reviews = {}
    for group_name, students in groups.items():
        #reviews_group[col[0]] is a team's review for a column
        reviews_group = {}
        reviews[group_name] = reviews_group
        for col in cols:
            # df each row is the student(first col name)'s review from other students(other col names)
            row = list(map(lambda x: f"{x[0] + ', ' + x[1]}", students))
            row.insert(0, col[0])
            row.append("Overall")#use this space to show the review column name
            df = pd.DataFrame(columns=row)
            reviews_group[col[0]] = df
            #none will not be counted in for mean later
            for student in students:
                r = [student[0] + ", " + student[1]]
                r.extend([None for s in students])#review from other students
                r.append(None)#overall
                df.loc[len(df)] = r
            df.set_index(col[0], inplace=True)
        for student in students:
            #find the xlsx file with the same name as the student
            choices = [s for s in xlsx_files if student[0] in s and student[1] in s]
            if len(choices)!=1:
                    print(f"Error: {student} has {len(choices)} submissions.")
            else:
                # print(f"Found {choices[0]} for {student}")
                review_to_self = False
                wb_submission = load_workbook(choices[0])
                ws_submission = wb_submission[wb_submission.sheetnames[0]]
                for row in ws_submission.iter_rows(min_row=2, max_col=len(cols)+1, max_row=len(students)+1):
                    row_values = []
                    for cell in row:
                        if cell.value == None:
                            print(f"Error: {student} has no name in {choices[0]}")
                            exit(-1)
                        elif cell.col_idx == 1:
                            name = cell.value
                            #do not take self review
                            review_to_self = (name == student[0]+", "+student[1])
                            if review_to_self: break
                        else:
                            row_values.append(cell.value)
                    if not review_to_self:
                        for i,col in enumerate(cols):
                            reviews_group[col[0]].loc[name,student[0]+", "+student[1]] = int(row_values[i])

        for c,col in enumerate(cols):
            df = reviews_group[col[0]]
            df["Overall"] = df.mean(axis=1)
            if os.path.exists(dst):
                with pd.ExcelWriter(dst,engine="openpyxl", mode="a",if_sheet_exists="overlay") as writer:
                    df.to_excel(writer, startrow=(c+1)*(len(students))+11+c*2, startcol=2, sheet_name=f"{group_name}")
            else:
                with pd.ExcelWriter(dst,engine="openpyxl") as writer:
                    df.to_excel(writer, startrow=len(students)+10, startcol=2, sheet_name=f"{group_name}")

        print("Complete collecting peer review files...")